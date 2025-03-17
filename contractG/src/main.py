#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
contractG
主程序入口
"""

import os
import sys
import locale
import traceback
from datetime import datetime
from openpyxl import load_workbook
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas

# 添加当前目录到Python路径
current_dir = os.path.dirname(os.path.abspath(__file__))
if current_dir not in sys.path:
    sys.path.insert(0, current_dir)

# 添加项目根目录到Python路径
project_root = os.path.dirname(current_dir)
if project_root not in sys.path:
    sys.path.insert(0, project_root)

# 导入路径设置模块
try:
    from src.utils.path_setup import setup_python_path
except ImportError:
    # 如果无法使用绝对导入，尝试相对导入
    try:
        from utils.path_setup import setup_python_path
    except ImportError:
        # 如果仍然失败，尝试直接导入
        import sys
        sys.path.append(os.path.join(current_dir, 'utils'))
        from path_setup import setup_python_path

def setup_encoding():
    """设置系统编码"""
    try:
        # 设置默认编码为UTF-8
        if sys.stdout.encoding != 'utf-8':
            sys.stdout.reconfigure(encoding='utf-8')
        if sys.stderr.encoding != 'utf-8':
            sys.stderr.reconfigure(encoding='utf-8')
            
        # 设置区域和语言
        locale.setlocale(locale.LC_ALL, 'zh_CN.UTF-8')
    except Exception:
        try:
            # 如果无法设置中文，则使用系统默认
            locale.setlocale(locale.LC_ALL, '')
        except Exception:
            pass

# 设置Python路径
setup_python_path()

# 导入UI相关模块
from PyQt5.QtWidgets import QApplication, QHBoxLayout, QPushButton
from PyQt5.QtCore import Qt, QTimer
from PyQt5.QtGui import QFont, QIcon

# 尝试使用不同的导入方式，以适应打包后的环境
try:
    # 首先尝试直接导入（开发环境）
    print("尝试导入方式1: from src.ui.main_window import MainWindow")
    from src.ui.main_window import MainWindow
    from src.ui.styles import GLOBAL_STYLE, FONT_FAMILY
    from src.ui.splash_screen import SplashScreen
    print("导入方式1成功")
except ImportError as e1:
    print(f"导入方式1失败: {e1}")
    try:
        # 如果失败，尝试相对导入（打包环境）
        print("尝试导入方式2: from ui.main_window import MainWindow")
        from ui.main_window import MainWindow
        from ui.styles import GLOBAL_STYLE, FONT_FAMILY
        from ui.splash_screen import SplashScreen
        print("导入方式2成功")
    except ImportError as e2:
        print(f"导入方式2失败: {e2}")
        try:
            # 如果还是失败，尝试直接导入
            print("尝试导入方式3: import main_window")
            import main_window
            import styles
            import splash_screen
            MainWindow = main_window.MainWindow
            GLOBAL_STYLE = styles.GLOBAL_STYLE
            FONT_FAMILY = styles.FONT_FAMILY
            SplashScreen = splash_screen.SplashScreen
            print("导入方式3成功")
        except ImportError as e3:
            print(f"导入方式3失败: {e3}")
            # 最后的尝试，动态加载模块
            print("尝试导入方式4: 动态加载模块")
            import importlib.util
            
            # 尝试不同的路径
            possible_paths = [
                os.path.join(current_dir, 'ui', 'main_window.py'),
                os.path.join(project_root, 'src', 'ui', 'main_window.py')
            ]
            
            if hasattr(sys, '_MEIPASS'):
                possible_paths.extend([
                    os.path.join(sys._MEIPASS, 'src', 'ui', 'main_window.py'),
                    os.path.join(sys._MEIPASS, 'ui', 'main_window.py')
                ])
            
            main_window_spec = None
            styles_spec = None
            splash_screen_spec = None
            
            for path in possible_paths:
                if os.path.exists(path):
                    print(f"找到main_window.py: {path}")
                    main_window_spec = importlib.util.spec_from_file_location("main_window", path)
                    styles_path = path.replace('main_window.py', 'styles.py')
                    splash_screen_path = path.replace('main_window.py', 'splash_screen.py')
                    if os.path.exists(styles_path):
                        print(f"找到styles.py: {styles_path}")
                        styles_spec = importlib.util.spec_from_file_location("styles", styles_path)
                    if os.path.exists(splash_screen_path):
                        print(f"找到splash_screen.py: {splash_screen_path}")
                        splash_screen_spec = importlib.util.spec_from_file_location("splash_screen", splash_screen_path)
                    break
            
            if main_window_spec and styles_spec and splash_screen_spec:
                main_window_module = importlib.util.module_from_spec(main_window_spec)
                styles_module = importlib.util.module_from_spec(styles_spec)
                splash_screen_module = importlib.util.module_from_spec(splash_screen_spec)
                main_window_spec.loader.exec_module(main_window_module)
                styles_spec.loader.exec_module(styles_module)
                splash_screen_spec.loader.exec_module(splash_screen_module)
                MainWindow = main_window_module.MainWindow
                GLOBAL_STYLE = styles_module.GLOBAL_STYLE
                FONT_FAMILY = styles_module.FONT_FAMILY
                SplashScreen = splash_screen_module.SplashScreen
                print("导入方式4成功")
            else:
                print("所有导入方式都失败，无法找到必要的模块")
                raise ImportError("无法导入必要的模块: main_window.py, styles.py 和 splash_screen.py")

def resource_path(relative_path):
    """获取资源文件的绝对路径"""
    try:
        # 检查是否在打包环境中运行
        if hasattr(sys, '_MEIPASS'):
            # 如果是打包环境，使用 _MEIPASS 作为基础路径
            base_path = sys._MEIPASS
        else:
            # 获取当前文件所在目录
            current_dir = os.path.dirname(os.path.abspath(__file__))
            # 获取项目根目录
            base_path = os.path.dirname(current_dir)
        
        # 构建并返回资源文件的完整路径
        full_path = os.path.join(base_path, 'resources', relative_path)
        
        # 确保目录存在
        os.makedirs(os.path.dirname(full_path), exist_ok=True)
        
        return full_path
    except Exception as e:
        print(f"Error in resource_path: {str(e)}")
        return relative_path

def config_path(relative_path):
    """获取配置文件的绝对路径"""
    try:
        # 检查是否在打包环境中运行
        if hasattr(sys, '_MEIPASS'):
            # 如果是打包环境，使用 _MEIPASS 作为基础路径
            base_path = sys._MEIPASS
        else:
            # 获取当前文件所在目录
            current_dir = os.path.dirname(os.path.abspath(__file__))
            # 获取项目根目录
            base_path = os.path.dirname(current_dir)
        
        # 构建并返回配置文件的完整路径
        full_path = os.path.join(base_path, 'config', relative_path)
        
        # 确保目录存在
        os.makedirs(os.path.dirname(full_path), exist_ok=True)
        
        return full_path
    except Exception as e:
        print(f"Error in config_path: {str(e)}")
        return relative_path

def ensure_directories():
    """确保必要的目录存在"""
    directories = [
        'data',
        'output',  # 输出目录
        'templates',  # 模板目录
        'logs',  # 日志目录
        'resources',  # 资源文件目录
        'config',  # 配置文件目录
        'src/database',  # 数据库目录
        'src/ui/dialogs',  # UI对话框目录
    ]
    for directory in directories:
        os.makedirs(directory, exist_ok=True)

def excel_to_pdf(excel_file, pdf_file):
    # 读取 Excel 文件
    wb = load_workbook(excel_file)
    ws = wb.active
    
    # 创建 PDF 文件
    c = canvas.Canvas(pdf_file, pagesize=A4)
    width, height = A4

    # 设置初始位置
    x_offset = 50
    y_offset = height - 50
    line_height = 20

    # 写入标题
    c.setFont("Helvetica-Bold", 12)
    c.drawString(x_offset, y_offset, "Excel to PDF Conversion")
    y_offset -= line_height

    # 写入列名
    c.setFont("Helvetica-Bold", 10)
    for col in ws.iter_cols(1, ws.max_column):
        c.drawString(x_offset, y_offset, str(col[0].value))
        x_offset += 100  # 调整列间距
    y_offset -= line_height
    x_offset = 50

    # 写入数据
    c.setFont("Helvetica", 10)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            c.drawString(x_offset, y_offset, str(cell.value))
            x_offset += 100
        y_offset -= line_height
        x_offset = 50
        if y_offset < 50:  # 如果接近页面底部，添加新页面
            c.showPage()
            y_offset = height - 50

    # 保存 PDF
    c.save()

def main():
    """主函数"""
    try:
        # 设置编码
        setup_encoding()
        
        # 确保必要的目录存在
        ensure_directories()
        
        # 创建QApplication实例
        app = QApplication(sys.argv)
        
        # 设置全局样式
        app.setStyleSheet(GLOBAL_STYLE)
        
        # 设置全局字体
        font = QFont(FONT_FAMILY, 9)
        app.setFont(font)
        
        # 设置应用程序图标
        icon_path = resource_path('icon.ico')
        if os.path.exists(icon_path):
            app.setWindowIcon(QIcon(icon_path))
        else:
            print(f"警告: 图标文件未找到: {icon_path}")
        
        # 显示启动画面
        splash = SplashScreen()
        splash.show()
        
        # 创建主窗口但不显示
        main_window = MainWindow()
        
        # 在启动画面完成后显示主窗口
        def finish_splash():
            splash.finish(main_window)
            main_window.show()
        
        # 设置定时器在3秒后关闭启动画面并显示主窗口
        QTimer.singleShot(3000, finish_splash)
        
        # 运行应用程序
        sys.exit(app.exec_())
        
    except Exception as e:
        print(f"程序启动失败: {str(e)}")
        traceback.print_exc()
        sys.exit(1)

if __name__ == "__main__":
    main() 