#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
商品管理标签页模块
"""

import os
import sys
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
from PyQt5.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QPushButton, 
    QTableWidget, QTableWidgetItem, QLineEdit, 
    QLabel, QDialog, QFormLayout, QMessageBox,
    QDoubleSpinBox, QComboBox, QFileDialog, QHeaderView,
    QGroupBox, QSplitter, QFrame, QSizePolicy
)
from PyQt5.QtCore import Qt, pyqtSignal, QSize, QTimer
from PyQt5.QtGui import QIcon, QFont

# 注意：路径设置已由path_setup模块处理，不需要在这里重复设置

from src.database.excel_manager import ExcelManager
from src.models.product import Product
from src.ui.styles import (
    PRIMARY_COLOR, SECONDARY_COLOR, SUCCESS_COLOR, 
    WARNING_COLOR, DANGER_COLOR, LIGHT_COLOR,
    HEADING_STYLE, SUBHEADING_STYLE, CARD_STYLE
)
from pypinyin import lazy_pinyin

class ProductDialog(QDialog):
    """商品信息对话框"""
    def __init__(self, product=None, parent=None):
        super().__init__(parent)
        self.product = product
        self.setWindowTitle("商品信息")
        self.setMinimumWidth(500)
        self.setMinimumHeight(400)
        self.setup_ui()
        if product:
            self.load_product_data()
    
    def setup_ui(self):
        """设置UI"""
        layout = QVBoxLayout()
        layout.setSpacing(15)
        
        # 添加标题
        title_label = QLabel("商品信息编辑")
        title_label.setStyleSheet(HEADING_STYLE)
        layout.addWidget(title_label)
        
        # 表单布局
        form_group = QGroupBox("商品基本信息")
        form_layout = QFormLayout()
        form_layout.setSpacing(10)
        form_layout.setLabelAlignment(Qt.AlignRight)
        
        # 商品名称
        self.name = QLineEdit()
        self.name.setMinimumHeight(30)
        self.name.setPlaceholderText("请输入商品名称（必填）")
        self.name.textChanged.connect(self.validate_input)
        
        # 规格型号
        self.model = QLineEdit()
        self.model.setMinimumHeight(30)
        self.model.setPlaceholderText("请输入规格型号（必填）")
        self.model.textChanged.connect(self.validate_input)
        
        # 单位
        self.unit = QComboBox()
        self.unit.addItems(["个", "台", "套", "件", "箱", "米", "千克", "吨", "其他"])
        self.unit.setEditable(True)
        self.unit.setMinimumHeight(30)
        self.unit.currentTextChanged.connect(self.validate_input)
        
        # 单价
        self.price = QDoubleSpinBox()
        self.price.setRange(0, 9999999.99)
        self.price.setDecimals(2)
        self.price.setSingleStep(1)
        self.price.setSuffix(" 元")
        self.price.setMinimumHeight(30)
        self.price.valueChanged.connect(self.validate_input)
        
        # 错误提示标签
        self.error_label = QLabel()
        self.error_label.setStyleSheet("color: red; font-size: 9pt;")
        self.error_label.setWordWrap(True)
        self.error_label.hide()
        
        form_layout.addRow("商品名称:", self.name)
        form_layout.addRow("规格型号:", self.model)
        form_layout.addRow("单位:", self.unit)
        form_layout.addRow("单价:", self.price)
        form_layout.addRow("", self.error_label)
        
        form_group.setLayout(form_layout)
        layout.addWidget(form_group)
        
        # 按钮布局
        button_layout = QHBoxLayout()
        button_layout.setSpacing(10)
        
        self.save_btn = QPushButton("保存")
        self.save_btn.clicked.connect(self.validate_and_accept)
        self.save_btn.setStyleSheet(f"background-color: {SUCCESS_COLOR};")
        self.save_btn.setMinimumHeight(36)
        self.save_btn.setMinimumWidth(100)
        self.save_btn.setEnabled(False)  # 初始状态禁用
        
        cancel_btn = QPushButton("取消")
        cancel_btn.clicked.connect(self.reject)
        cancel_btn.setMinimumHeight(36)
        cancel_btn.setMinimumWidth(100)
        
        button_layout.addStretch()
        button_layout.addWidget(self.save_btn)
        button_layout.addWidget(cancel_btn)
        
        layout.addLayout(button_layout)
        self.setLayout(layout)
    
    def validate_input(self):
        """验证输入"""
        errors = []
        
        # 验证商品名称
        name = self.name.text().strip()
        if not name:
            errors.append("商品名称不能为空")
        elif len(name) > 100:
            errors.append("商品名称不能超过100个字符")
        
        # 验证规格型号
        model = self.model.text().strip()
        if not model:
            errors.append("规格型号不能为空")
        elif len(model) > 50:
            errors.append("规格型号不能超过50个字符")
        
        # 验证单位
        unit = self.unit.currentText().strip()
        if not unit:
            errors.append("单位不能为空")
        elif len(unit) > 10:
            errors.append("单位不能超过10个字符")
        
        # 验证单价
        price = self.price.value()
        if price <= 0:
            errors.append("单价必须大于0")
        elif price > 9999999.99:
            errors.append("单价不能超过9999999.99")
        
        # 显示或隐藏错误信息
        if errors:
            self.error_label.setText("• " + "\n• ".join(errors))
            self.error_label.show()
            self.save_btn.setEnabled(False)
        else:
            self.error_label.hide()
            self.save_btn.setEnabled(True)
    
    def validate_and_accept(self):
        """验证并接受"""
        self.validate_input()
        if self.save_btn.isEnabled():
            self.accept()
    
    def load_product_data(self):
        """加载商品数据"""
        if self.product:
            self.name.setText(self.product.name)
            self.model.setText(self.product.model)
            
            # 设置单位
            index = self.unit.findText(self.product.unit)
            if index >= 0:
                self.unit.setCurrentIndex(index)
            else:
                self.unit.setCurrentText(self.product.unit)
            
            self.price.setValue(float(self.product.price))
    
    def get_product_data(self):
        """获取商品数据"""
        product = Product(
            name=self.name.text(),
            model=self.model.text(),
            unit=self.unit.currentText(),
            price=str(self.price.value())
        )
        return product


class ProductTab(QWidget):
    """商品管理标签页"""
    product_updated = pyqtSignal()
    
    def __init__(self):
        super().__init__()
        self.excel_manager = ExcelManager()
        self.products = []
        self.filtered_products = []  # 缓存过滤后的商品列表
        self.search_timer = QTimer()  # 添加定时器用于防抖
        self.search_timer.setSingleShot(True)
        self.search_timer.timeout.connect(self._do_filter_products)
        self.setup_ui()
        self.load_products()
    
    def setup_ui(self):
        """设置UI"""
        layout = QVBoxLayout()
        layout.setSpacing(15)  # 增加间距
        
        # 添加标题和说明
        header_layout = QHBoxLayout()
        title_label = QLabel("商品管理")
        title_label.setStyleSheet(HEADING_STYLE)
        header_layout.addWidget(title_label)
        
        # 添加操作按钮到标题行
        header_layout.addStretch()
        
        # 导入导出按钮
        import_btn = QPushButton("导入商品")
        import_btn.clicked.connect(self.import_products)
        import_btn.setStyleSheet(f"background-color: {SECONDARY_COLOR};")
        import_btn.setMinimumHeight(32)  # 增加按钮高度
        header_layout.addWidget(import_btn)
        
        export_btn = QPushButton("导出商品")
        export_btn.clicked.connect(self.export_products)
        export_btn.setStyleSheet(f"background-color: {SECONDARY_COLOR};")
        export_btn.setMinimumHeight(32)
        header_layout.addWidget(export_btn)
        
        layout.addLayout(header_layout)
        
        # 添加说明文字
        desc_label = QLabel("在此管理商品信息，可以添加、编辑、删除商品，也可以导入导出商品数据。")
        desc_label.setStyleSheet("color: #666; font-size: 9pt;")
        layout.addWidget(desc_label)
        
        # 搜索框
        search_layout = QHBoxLayout()
        search_layout.setSpacing(10)  # 增加间距
        
        search_label = QLabel("搜索:")
        search_label.setMinimumWidth(50)
        search_layout.addWidget(search_label)
        
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("输入商品名称或规格型号进行搜索...")
        self.search_input.textChanged.connect(self.filter_products)
        self.search_input.setMinimumHeight(32)  # 增加搜索框高度
        search_layout.addWidget(self.search_input)
        
        layout.addLayout(search_layout)
        
        # 商品表格
        self.product_table = QTableWidget()
        self.product_table.setColumnCount(4)
        self.product_table.setHorizontalHeaderLabels(["商品名称", "规格型号", "单位", "单价(元)"])
        self.product_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.product_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.product_table.doubleClicked.connect(self.edit_product)
        
        # 设置表格列宽
        self.product_table.setColumnWidth(0, 300)  # 商品名称
        self.product_table.setColumnWidth(1, 200)  # 规格型号
        self.product_table.setColumnWidth(2, 100)  # 单位
        self.product_table.setColumnWidth(3, 150)  # 单价
        
        # 设置表格样式
        self.product_table.horizontalHeader().setStretchLastSection(True)
        self.product_table.horizontalHeader().setMinimumHeight(40)  # 增加表头高度
        self.product_table.horizontalHeader().setStyleSheet("QHeaderView::section { padding: 6px; }")
        self.product_table.verticalHeader().setVisible(False)
        self.product_table.setAlternatingRowColors(True)
        self.product_table.setStyleSheet("""
            alternate-background-color: #F5F5F5;
            QTableWidget {
                gridline-color: #E0E0E0;
                selection-background-color: #E3F2FD;
                selection-color: #212121;
            }
            QTableWidget::item {
                padding: 8px;
                min-height: 28px;
            }
        """)
        
        # 设置表格占据大部分空间
        self.product_table.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        layout.addWidget(self.product_table, 1)  # 添加拉伸因子
        
        # 按钮布局
        button_layout = QHBoxLayout()
        button_layout.setSpacing(10)  # 增加按钮间距
        
        add_btn = QPushButton("添加商品")
        add_btn.clicked.connect(self.add_product)
        add_btn.setStyleSheet(f"background-color: {SUCCESS_COLOR};")
        add_btn.setMinimumHeight(36)  # 增加按钮高度
        add_btn.setMinimumWidth(120)  # 增加按钮宽度
        button_layout.addWidget(add_btn)
        
        edit_btn = QPushButton("编辑商品")
        edit_btn.clicked.connect(self.edit_product)
        edit_btn.setMinimumHeight(36)
        edit_btn.setMinimumWidth(120)
        button_layout.addWidget(edit_btn)
        
        delete_btn = QPushButton("删除商品")
        delete_btn.clicked.connect(self.delete_product)
        delete_btn.setStyleSheet(f"background-color: {DANGER_COLOR};")
        delete_btn.setMinimumHeight(36)
        delete_btn.setMinimumWidth(120)
        button_layout.addWidget(delete_btn)
        
        button_layout.addStretch()
        
        refresh_btn = QPushButton("刷新列表")
        refresh_btn.clicked.connect(self.load_products)
        refresh_btn.setMinimumHeight(36)
        refresh_btn.setMinimumWidth(120)
        button_layout.addWidget(refresh_btn)
        
        layout.addLayout(button_layout)
        
        self.setLayout(layout)
    
    def load_products(self):
        """加载商品数据"""
        self.products = self.excel_manager.load_products()
        self.update_table()
    
    def update_table(self, products_to_display=None):
        """更新表格显示"""
        if products_to_display is None:
            products_to_display = self.products
            
        self.product_table.setRowCount(0)
        
        # 批量添加行以提高性能
        self.product_table.setRowCount(len(products_to_display))
        
        for i, product in enumerate(products_to_display):
            self.product_table.setItem(i, 0, QTableWidgetItem(str(product.name)))
            self.product_table.setItem(i, 1, QTableWidgetItem(str(product.model)))
            self.product_table.setItem(i, 2, QTableWidgetItem(str(product.unit)))
            self.product_table.setItem(i, 3, QTableWidgetItem(str(product.price)))
    
    def filter_products(self):
        """触发商品过滤"""
        # 重置定时器，实现防抖
        self.search_timer.stop()
        self.search_timer.start(300)  # 300ms 后执行实际的过滤操作
    
    def _do_filter_products(self):
        """实际执行商品过滤"""
        search_text = self.search_input.text().lower()
        
        # 如果搜索框为空，显示所有商品
        if not search_text:
            self.filtered_products = self.products
            self.update_table(self.filtered_products)
            return
        
        # 将搜索关键词按空格分割，支持多个关键词
        search_keywords = [keyword.strip() for keyword in search_text.split() if keyword.strip()]
        
        # 过滤商品
        self.filtered_products = []
        for product in self.products:
            # 获取商品信息的拼音和首字母
            name_pinyin = ''.join(lazy_pinyin(product.name))
            name_initials = ''.join([p[0] for p in lazy_pinyin(product.name)])
            model_pinyin = ''.join(lazy_pinyin(product.model))
            model_initials = ''.join([p[0] for p in lazy_pinyin(product.model)])
            
            # 将所有可搜索字段组合成一个字符串
            searchable_text = f"{product.name} {product.model} {product.unit} {name_pinyin} {name_initials} {model_pinyin} {model_initials}".lower()
            
            # 检查是否所有关键词都匹配
            if all(keyword in searchable_text for keyword in search_keywords):
                self.filtered_products.append(product)
        
        # 更新表格显示
        self.update_table(self.filtered_products)
    
    def add_product(self):
        """添加商品"""
        dialog = ProductDialog(parent=self)
        if dialog.exec_():
            product = dialog.get_product_data()
            self.products.append(product)
            self.excel_manager.save_products(self.products)
            self.update_table()
            self.product_updated.emit()
            QMessageBox.information(self, "成功", "商品添加成功！")
    
    def edit_product(self):
        """编辑商品"""
        selected_rows = self.product_table.selectedIndexes()
        if not selected_rows:
            QMessageBox.warning(self, "警告", "请先选择一个商品！")
            return
        
        row = selected_rows[0].row()
        product = self.products[row]
        
        dialog = ProductDialog(product, self)
        if dialog.exec_():
            updated_product = dialog.get_product_data()
            self.products[row] = updated_product
            self.excel_manager.save_products(self.products)
            self.update_table()
            self.product_updated.emit()
            QMessageBox.information(self, "成功", "商品信息更新成功！")
    
    def delete_product(self):
        """删除商品"""
        selected_rows = self.product_table.selectedIndexes()
        if not selected_rows:
            QMessageBox.warning(self, "警告", "请先选择一个商品！")
            return
        
        row = selected_rows[0].row()
        product = self.products[row]
        
        reply = QMessageBox.question(
            self, 
            "确认删除", 
            f"确定要删除商品 '{product.name}' 吗？",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )
        
        if reply == QMessageBox.Yes:
            del self.products[row]
            self.excel_manager.save_products(self.products)
            self.update_table()
            self.product_updated.emit()
            QMessageBox.information(self, "成功", "商品删除成功！")
    
    def get_products(self):
        """获取所有商品"""
        return self.products
    
    def import_products(self):
        """导入商品数据"""
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "选择要导入的Excel文件",
            os.getcwd(),
            "Excel Files (*.xlsx)"
        )
        
        if not file_path:
            return
        
        # 读取Excel文件
        try:
            wb = load_workbook(file_path)
            ws = wb.active
            
            # 获取表头
            headers = [str(cell.value).strip() if cell.value else '' for cell in ws[1]]
            
            # 列名映射（支持多种可能的列名）
            column_mapping = {
                'name': ['商品名称', '名称', 'name', '产品名称'],
                'model': ['规格型号', '型号', 'model', '规格'],
                'unit': ['单位', 'unit', '计量单位'],
                'price': ['单价(元)', '单价', '价格', 'price', '单价（元）']
            }
            
            # 为每个必需字段找到对应的列索引
            header_indices = {}
            missing_fields = []
            
            for field, possible_names in column_mapping.items():
                found = False
                for name in possible_names:
                    if name in headers:
                        header_indices[field] = headers.index(name)
                        found = True
                        break
                if not found:
                    missing_fields.append(f"{possible_names[0]}({field})")
            
            if missing_fields:
                # 创建模板文件
                template_path = os.path.join(os.path.dirname(file_path), "商品导入模板.xlsx")
                template_wb = Workbook()
                template_ws = template_wb.active
                template_ws.append([column_mapping[field][0] for field in ['name', 'model', 'unit', 'price']])
                
                # 添加示例数据
                template_ws.append(['示例商品1', 'XH-001', '个', '100.00'])
                template_ws.append(['示例商品2', 'XH-002', '件', '200.00'])
                
                # 设置列宽和样式
                template_ws.column_dimensions['A'].width = 40  # 商品名称
                template_ws.column_dimensions['B'].width = 30  # 规格型号
                template_ws.column_dimensions['C'].width = 15  # 单位
                template_ws.column_dimensions['D'].width = 15  # 单价
                
                # 设置表头样式
                header_font = Font(bold=True)
                for cell in template_ws[1]:
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center')
                
                template_wb.save(template_path)
                
                QMessageBox.warning(
                    self,
                    "格式错误",
                    f"Excel文件缺少必要的列：{', '.join(missing_fields)}\n"
                    f"已在同目录下创建模板文件：{os.path.basename(template_path)}\n"
                    "请使用正确的模板文件重新导入！"
                )
                return
            
            # 导入商品数据
            new_products = []
            validation_errors = []
            
            # 从第二行开始读取数据
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
                try:
                    # 获取数据
                    name = str(row[header_indices['name']] or '').strip()
                    model = str(row[header_indices['model']] or '').strip()
                    unit = str(row[header_indices['unit']] or '').strip()
                    price = str(row[header_indices['price']] or '0.00').strip()
                    
                    # 验证数据
                    if not name:
                        validation_errors.append(f"第 {row_idx} 行：商品名称不能为空")
                        continue
                    
                    try:
                        price_float = float(price.replace(',', ''))  # 处理可能的千位分隔符
                        if price_float < 0:
                            validation_errors.append(f"第 {row_idx} 行：单价不能为负数")
                            continue
                        price = f"{price_float:.2f}"  # 统一格式化为两位小数
                    except ValueError:
                        validation_errors.append(f"第 {row_idx} 行：单价必须是数字")
                        continue
                    
                    # 创建商品对象
                    product = Product(
                        name=name,
                        model=model,
                        unit=unit,
                        price=price
                    )
                    new_products.append(product)
                
                except Exception as e:
                    validation_errors.append(f"第 {row_idx} 行：数据格式错误 - {str(e)}")
            
            # 如果有验证错误
            if validation_errors:
                error_msg = "\n".join(validation_errors)
                QMessageBox.warning(self, "导入警告", f"部分数据导入失败：\n{error_msg}")
            
            # 添加有效的商品数据
            if new_products:
                self.products.extend(new_products)
                self.save_products()
                self.load_products()
                QMessageBox.information(
                    self, 
                    "导入成功", 
                    f"成功导入 {len(new_products)} 个商品"
                )
        
        except Exception as e:
            QMessageBox.critical(self, "错误", f"导入商品数据失败：{str(e)}")

    def export_products(self):
        """导出商品数据"""
        if not self.products:
            QMessageBox.warning(self, "警告", "没有商品数据可导出")
            return
        
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "保存商品数据",
            os.path.join(os.getcwd(), "商品数据.xlsx"),
            "Excel Files (*.xlsx)"
        )
        
        if not file_path:
            return
        
        try:
            # 创建工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = '商品数据'
            
            # 添加表头
            headers = ['商品名称', '规格型号', '单位', '单价(元)']
            ws.append(headers)
            
            # 设置表头样式
            header_font = Font(bold=True)
            for cell in ws[1]:
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
            
            # 添加数据
            for product in self.products:
                row = [
                    product.name,
                    product.model,
                    product.unit,
                    float(product.price)
                ]
                ws.append(row)
            
            # 调整列宽
            ws.column_dimensions['A'].width = 40  # 商品名称
            ws.column_dimensions['B'].width = 30  # 规格型号
            ws.column_dimensions['C'].width = 15  # 单位
            ws.column_dimensions['D'].width = 15  # 单价
            
            # 保存文件
            wb.save(file_path)
            
            QMessageBox.information(
                self, 
                "导出成功", 
                f"成功导出 {len(self.products)} 个商品到：\n{file_path}"
            )
        
        except Exception as e:
            QMessageBox.critical(self, "错误", f"导出商品数据失败：{str(e)}") 

    def save_products(self, products=None):
        """保存商品数据到Excel文件"""
        try:
            # 如果没有提供products参数，使用当前的products列表
            products_to_save = products if products is not None else self.products
            
            # 使用excel_manager保存数据
            self.excel_manager.save_products(products_to_save)
            
            # 更新当前products列表
            if products is not None:
                self.products = products
            
            # 更新表格显示
            self.update_table()
            
            # 发送更新信号
            self.product_updated.emit()
            
            return True
        except Exception as e:
            QMessageBox.critical(self, "错误", f"保存商品数据失败：{str(e)}")
            return False