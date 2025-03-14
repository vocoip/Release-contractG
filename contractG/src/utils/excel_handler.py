#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
Excel处理工具模块
"""

import os
import datetime
import random
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, Color
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
from openpyxl.cell.rich_text import TextBlock, InlineFont
from PyQt5.QtWidgets import QMessageBox
import win32com.client
from PyPDF2 import PdfReader, PdfWriter
from PIL import Image as PILImage
import io
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.lib.utils import ImageReader
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import time
import threading
import shutil
import tempfile
import sys
from pathlib import Path
import pythoncom
import fitz  # PyMuPDF
import logging

# 导入PyMuPDF库
try:
    import fitz  # PyMuPDF
    PYMUPDF_AVAILABLE = True
except ImportError:
    print("警告: PyMuPDF库未安装，将不会生成图片式PDF。请运行 'pip install PyMuPDF' 安装。")
    PYMUPDF_AVAILABLE = False

# 注册中文字体
try:
    # 尝试注册微软雅黑字体（Windows系统常见字体）
    pdfmetrics.registerFont(TTFont('SimSun', 'C:/Windows/Fonts/simsun.ttc'))
except:
    try:
        # 尝试注册宋体字体（Windows系统常见字体）
        pdfmetrics.registerFont(TTFont('SimSun', 'C:/Windows/Fonts/simhei.ttf'))
    except:
        # 如果都失败，使用Helvetica（ReportLab默认支持的字体）
        pass

class ExcelHandler:
    """Excel处理工具类"""
    def __init__(self):
        # 确保输出目录存在
        self.output_dir = os.path.join('output', 'contracts')
        os.makedirs(self.output_dir, exist_ok=True)
        
        # 设置日志
        self.logger = self._setup_logger()
        
        # 初始化回调函数
        self.log_callback = None
    
    def set_log_callback(self, callback):
        """设置日志回调函数，用于将日志信息传递给UI"""
        self.log_callback = callback
    
    def log(self, message):
        """记录日志信息，并通过回调函数传递给UI"""
        # 记录到日志文件
        self.logger.info(message)
        
        # 如果设置了回调函数，则调用回调函数
        if self.log_callback:
            self.log_callback(message)
        else:
            # 如果没有设置回调函数，则打印到控制台
            print(message)
    
    def _setup_logger(self):
        """设置日志记录器"""
        logger = logging.getLogger('ExcelHandler')
        logger.setLevel(logging.INFO)
        
        # 确保日志目录存在
        log_dir = 'logs'
        os.makedirs(log_dir, exist_ok=True)
        
        # 创建文件处理器
        log_file = os.path.join(log_dir, 'excel_handler.log')
        file_handler = logging.FileHandler(log_file, encoding='utf-8')
        file_handler.setLevel(logging.INFO)
        
        # 创建格式化器
        formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
        file_handler.setFormatter(formatter)
        
        # 添加处理器到记录器
        logger.addHandler(file_handler)
        
        return logger
    
    def get_contracts_folder(self):
        """获取合同文件夹路径，用于前端打开文件夹按钮"""
        return os.path.abspath(self.output_dir)
    
    def generate_contract(self, contract):
        """生成合同文件"""
        # 为合同编号添加三位随机数
        if not hasattr(contract, 'random_suffix') or not contract.random_suffix:
            contract.random_suffix = f"{random.randint(100, 999)}"
        
        # 生成合同和报价单
        contract_file = self._generate_contract_file(contract)
        quote_file = self._generate_quote_file(contract)
        
        # 如果需要添加印章
        if hasattr(contract, 'add_seal') and contract.add_seal:
            # 将Excel转换为PDF
            contract_pdf = self.convert_excel_to_pdf(contract_file)
            
            # 确保PDF为A4尺寸
            contract_pdf = self.ensure_a4_size_pdf(contract_pdf)
            
            # 添加印章
            seal_file = self._get_seal_file(contract)
            if seal_file and os.path.exists(seal_file):
                # 获取印章位置
                position = 'right-bottom'
                if hasattr(contract, 'seal_position'):
                    position = contract.seal_position
                
                # 获取签章文字
                seal_text = None
                if hasattr(contract, 'seal_text'):
                    seal_text = contract.seal_text
                
                # 添加印章到PDF
                contract_pdf_with_seal = self.add_seal_to_pdf(
                    contract_pdf, 
                    seal_file, 
                    position,
                    seal_text=seal_text
                )
                
                # 如果需要转换为图片式PDF
                if hasattr(contract, 'convert_to_image_pdf') and contract.convert_to_image_pdf:
                    # 使用改进的方法转换为图片式PDF，确保内容完整
                    self.log(f"正在将合同PDF转换为图片式PDF，确保内容完整...")
                    contract_pdf_final = self.convert_pdf_to_image_pdf(
                        contract_pdf_with_seal, 
                        dpi=200,  # 使用更高的DPI确保内容清晰
                        quality=85  # 使用更高的质量确保内容清晰
                    )
                    self.log(f"合同图片式PDF转换完成: {contract_pdf_final}")
                    
                    # 删除中间生成的PDF文件
                    if os.path.exists(contract_pdf_with_seal) and contract_pdf_with_seal != contract_pdf_final:
                        os.remove(contract_pdf_with_seal)
                        self.log(f"已删除中间PDF文件: {contract_pdf_with_seal}")
                    
                    return contract_file, quote_file, contract_pdf_final
                
                return contract_file, quote_file, contract_pdf_with_seal
            
            return contract_file, quote_file, contract_pdf
        
        return contract_file, quote_file
    
    def _get_seal_file(self, contract):
        """获取印章文件路径"""
        # 默认印章文件
        seal_file = os.path.join('resources', 'seals', '392328178266522.png')
        
        # 如果合同对象指定了印章文件
        if hasattr(contract, 'seal_file') and contract.seal_file:
            seal_file = contract.seal_file
        
        # 如果公司对象指定了印章文件
        elif hasattr(contract.company, 'seal_file') and contract.company.get('seal_file'):
            seal_file = contract.company.get('seal_file')
        
        # 检查文件是否存在
        if not os.path.exists(seal_file):
            self.logger.warning(f"印章文件不存在: {seal_file}")
            return None
        
        return seal_file
    
    def convert_excel_to_pdf(self, excel_file):
        """
        将Excel文件转换为PDF
        
        参数:
            excel_file (str): Excel文件路径
            
        返回:
            str: 生成的PDF文件路径
        """
        # 获取输出PDF文件路径
        pdf_file = os.path.splitext(excel_file)[0] + '.pdf'
        
        self.log(f"开始将Excel转换为PDF: {os.path.basename(excel_file)}")
        
        try:
            # 初始化COM组件
            pythoncom.CoInitialize()
            
            # 创建Excel应用程序实例
            excel = win32com.client.Dispatch("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            
            # 打开Excel文件
            workbook = excel.Workbooks.Open(os.path.abspath(excel_file))
            
            # 转换为PDF
            workbook.ExportAsFixedFormat(
                Type=0,  # 0表示PDF格式
                Filename=os.path.abspath(pdf_file),
                Quality=0,  # 标准质量
                IncludeDocProperties=True,
                IgnorePrintAreas=False,
                OpenAfterPublish=False
            )
            
            # 关闭工作簿和Excel应用程序
            workbook.Close(False)
            excel.Quit()
            
            # 释放COM对象
            del workbook
            del excel
            
            # 释放COM组件
            pythoncom.CoUninitialize()
            
            self.log(f"Excel转换为PDF成功: {os.path.basename(pdf_file)}")
            return pdf_file
            
        except Exception as e:
            self.log(f"Excel转换为PDF失败: {str(e)}")
            raise
    
    def add_seal_to_pdf(self, pdf_file, image_file, position='right-bottom', output_pdf=None, seal_text=None):
        """
        向PDF文件添加图片（如印章）和文字
        
        参数:
            pdf_file (str): PDF文件路径
            image_file (str): 图片文件路径
            position (str): 图片位置，可选值：'right-bottom', 'right-top', 'left-bottom', 'left-top', 'center'
            output_pdf (str, optional): 输出PDF文件路径，如果为None则覆盖原文件
            seal_text (str, optional): 印章位置上方显示的文字
            
        返回:
            str: 生成的PDF文件路径
        """
        # 如果输出文件路径为None，则使用临时文件，之后再替换原文件
        is_same_file = False
        if output_pdf is None:
            is_same_file = True
            output_pdf = os.path.splitext(pdf_file)[0] + "_temp.pdf"
        
        self.log(f"开始向PDF添加印章: {os.path.basename(pdf_file)}")
        
        try:
            # 打开PDF文件
            pdf_document = fitz.open(pdf_file)
            
            # A4纸张尺寸（点，1英寸=72点）
            # A4尺寸为210mm x 297mm，转换为点为595 x 842
            a4_width = 595
            a4_height = 842
            
            # 打开图片
            img = fitz.open(image_file)
            img_rect = img[0].rect
            
            # 调整印章尺寸 - 设置印章宽度为A4宽度的20%，高度按比例缩放
            # 这样印章大小会更合适，不会太大也不会太小
            desired_width = a4_width * 0.2  # A4宽度的20%
            scale_factor = desired_width / img_rect.width
            new_img_width = img_rect.width * scale_factor
            new_img_height = img_rect.height * scale_factor
            
            self.log(f"调整印章尺寸: {img_rect.width:.1f}x{img_rect.height:.1f} -> {new_img_width:.1f}x{new_img_height:.1f}")
            
            # 创建一个新的PDF文档用于输出
            output_document = fitz.open()
            
            # 复制每一页并添加印章
            for page_number in range(pdf_document.page_count):
                # 获取原始页面
                original_page = pdf_document[page_number]
                
                # 创建新页面
                new_page = output_document.new_page(
                    width=original_page.rect.width,
                    height=original_page.rect.height
                )
                
                # 复制原始内容到新页面
                new_page.show_pdf_page(
                    new_page.rect,
                    pdf_document,
                    page_number
                )
                
                # 计算图片位置，使用调整后的尺寸
                if position == 'right-bottom':
                    x = a4_width - new_img_width - 80  # 右边距80点
                    y = a4_height - new_img_height - 80  # 下边距80点
                elif position == 'right-top':
                    x = a4_width - new_img_width - 80
                    y = 80
                elif position == 'left-bottom':
                    x = 80
                    y = a4_height - new_img_height - 80
                elif position == 'left-top':
                    x = 80
                    y = 80
                elif position == 'center':
                    x = (a4_width - new_img_width) / 2
                    y = (a4_height - new_img_height) / 2
                else:
                    # 默认右下角
                    x = a4_width - new_img_width - 80
                    y = a4_height - new_img_height - 80
                
                # 如果需要添加文字
                if seal_text:
                    # 设置文字样式
                    text_font_size = 12
                    text_color = (0, 0, 0)  # 黑色
                    
                    # 创建一个临时的文本对象来计算文字宽度
                    text_width = fitz.get_text_length(seal_text, fontname="china-ss", fontsize=text_font_size)
                    
                    # 计算文字位置（在印章上方20点，水平居中）
                    text_x = x + (new_img_width - text_width) / 2  # 文字水平居中于印章
                    text_y = y + 65  # 文字在印章上方20点
                    
                    # 创建文字对象
                    new_page.insert_text(
                        point=(text_x, text_y),
                        text=seal_text,
                        fontname="china-ss",  # 使用中文字体
                        fontsize=text_font_size,
                        color=text_color
                    )
                
                # 创建一个矩形区域来放置图片，使用调整后的尺寸
                seal_rect = fitz.Rect(x, y, x + new_img_width, y + new_img_height)
                
                # 将图片插入到页面
                new_page.insert_image(seal_rect, filename=image_file)
            
            # 保存修改后的PDF
            output_document.save(output_pdf)
            output_document.close()
            pdf_document.close()
            
            # 如果是覆盖原文件，则在保存完成后替换原文件
            if is_same_file:
                # 确保关闭文件后再替换
                if os.path.exists(pdf_file):
                    os.remove(pdf_file)
                os.rename(output_pdf, pdf_file)
                output_pdf = pdf_file
            
            self.log(f"成功向PDF添加印章: {os.path.basename(output_pdf)}")
            return output_pdf
            
        except Exception as e:
            self.log(f"向PDF添加印章时出错: {str(e)}")
            raise
    
    def convert_pdf_to_image_pdf(self, pdf_file, output_pdf=None, dpi=150, quality=60):
        """
        将PDF转换为图片式PDF（每页转为图片再嵌入新PDF）
        这样可以确保印章等元素在任何PDF阅读器中都能正确显示
        
        参数:
            pdf_file (str): PDF文件路径
            output_pdf (str, optional): 输出PDF文件路径，如果为None则使用原文件名加后缀
            dpi (int): 图片分辨率，默认150（降低以减小文件大小）
            quality (int): 图片压缩质量，范围1-100，默认60（降低以减小文件大小）
            
        返回:
            str: 生成的图片式PDF文件路径
        """
        # 如果输出文件路径为None，则使用默认的输出路径
        is_same_file = False
        if output_pdf is None:
            # 使用原文件名，但添加_image后缀
            output_pdf = os.path.splitext(pdf_file)[0] + '_image.pdf'
        elif output_pdf == pdf_file:
            # 如果输出文件与输入文件相同，使用临时文件
            is_same_file = True
            output_pdf = os.path.splitext(pdf_file)[0] + '_temp_image.pdf'
        
        self.log(f"开始将PDF转换为图片式PDF: {os.path.basename(pdf_file)}")
        
        try:
            # 打开PDF文件
            pdf_document = fitz.open(pdf_file)
            
            # 创建一个新的PDF文档
            output_document = fitz.open()
            
            # A4纸张尺寸（点，1英寸=72点）
            # A4尺寸为210mm x 297mm，转换为点为595 x 842
            a4_width = 595
            a4_height = 842
            
            # 遍历所有页面
            for page_num in range(len(pdf_document)):
                page = pdf_document[page_num]
                
                # 获取原始页面尺寸
                orig_width = page.rect.width
                orig_height = page.rect.height
                
                # 使用更高的DPI来确保内容清晰
                render_dpi = max(dpi, 200)  # 确保至少200 DPI以保证内容清晰
                
                self.log(f"处理第{page_num+1}页，尺寸: {orig_width:.1f}x{orig_height:.1f}，DPI: {render_dpi}")
                
                # 将页面渲染为图片，使用较高的DPI以确保内容完整
                # 不应用缩放，直接使用原始尺寸渲染，避免内容丢失
                pix = page.get_pixmap(matrix=fitz.Matrix(render_dpi/72, render_dpi/72))
                
                # 创建临时图像文件以应用压缩
                temp_img_path = os.path.join(tempfile.gettempdir(), f"temp_page_{page_num}.jpg")
                
                # 保存为JPEG格式并应用压缩，但使用较高的质量以确保内容清晰
                pix.save(temp_img_path, output="jpeg", jpg_quality=max(quality, 80))
                
                # 创建一个新页面，使用A4尺寸
                new_page = output_document.new_page(width=a4_width, height=a4_height)
                
                # 加载保存的图像以获取其尺寸
                img = PILImage.open(temp_img_path)
                img_width, img_height = img.size
                img.close()
                
                # 计算缩放比例，确保图像适合A4尺寸并且内容完整
                scale_x = a4_width / img_width
                scale_y = a4_height / img_height
                scale = min(scale_x, scale_y) * 0.95  # 使用95%的缩放比例，留出边距
                
                # 计算图像在A4页面上的位置（居中）
                scaled_width = img_width * scale
                scaled_height = img_height * scale
                center_x = (a4_width - scaled_width) / 2
                center_y = (a4_height - scaled_height) / 2
                
                # 确保图像不会超出A4页面
                if center_x < 0:
                    center_x = 0
                if center_y < 0:
                    center_y = 0
                
                # 创建目标矩形
                target_rect = fitz.Rect(
                    center_x, center_y, 
                    center_x + scaled_width, center_y + scaled_height
                )
                
                # 将图片插入到新页面
                new_page.insert_image(target_rect, filename=temp_img_path)
                
                # 删除临时文件
                try:
                    os.remove(temp_img_path)
                except:
                    pass
            
            # 保存修改后的PDF
            output_document.save(output_pdf)
            output_document.close()
            pdf_document.close()
            
            # 如果是覆盖原文件，则在保存完成后替换原文件
            if is_same_file:
                # 确保关闭文件后再替换
                if os.path.exists(pdf_file):
                    os.remove(pdf_file)
                os.rename(output_pdf, pdf_file)
                output_pdf = pdf_file
            
            # 检查文件大小，如果太大则尝试进一步压缩
            file_size_mb = os.path.getsize(output_pdf) / (1024 * 1024)
            if file_size_mb > 2 and dpi > 120 and quality > 50:
                self.log(f"文件大小超过2MB ({file_size_mb:.2f}MB)，尝试进一步压缩...")
                # 递归调用自身，降低DPI和质量，但保持最低限度以确保内容清晰
                return self.convert_pdf_to_image_pdf(pdf_file, output_pdf, dpi=120, quality=70)
            
            self.log(f"成功将PDF转换为图片式PDF: {os.path.basename(output_pdf)}")
            return output_pdf
            
        except Exception as e:
            self.log(f"将PDF转换为图片式PDF时出错: {str(e)}")
            raise
    
    def generate_contract_only(self, contract):
        """仅生成合同文件"""
        # 为合同编号添加三位随机数
        if not hasattr(contract, 'random_suffix') or not contract.random_suffix:
            contract.random_suffix = f"{random.randint(100, 999)}"
        
        # 仅生成合同
        self.log(f"开始生成合同文件: {contract.number}")
        contract_file = self._generate_contract_file(contract)
        self.log(f"合同Excel文件已生成: {os.path.basename(contract_file)}")
        
        # 如果需要添加印章
        if hasattr(contract, 'convert_to_pdf') and contract.convert_to_pdf:
            # 将Excel转换为PDF
            contract_pdf = self.convert_excel_to_pdf(contract_file)
            
            # 确保PDF为A4尺寸
            contract_pdf = self.ensure_a4_size_pdf(contract_pdf)
            
            # 如果需要添加印章
            if hasattr(contract, 'add_seal') and contract.add_seal:
                # 添加印章
                seal_file = self._get_seal_file(contract)
                if seal_file and os.path.exists(seal_file):
                    # 获取印章位置
                    position = 'right-bottom'
                    if hasattr(contract, 'seal_position'):
                        position = contract.seal_position
                    
                    # 获取签章文字
                    seal_text = None
                    if hasattr(contract, 'seal_text'):
                        seal_text = contract.seal_text
                    
                    # 添加印章到PDF
                    contract_pdf_with_seal = self.add_seal_to_pdf(
                        contract_pdf, 
                        seal_file, 
                        position,
                        seal_text=seal_text
                    )
                    
                    # 如果需要转换为图片式PDF
                    if hasattr(contract, 'convert_to_image_pdf') and contract.convert_to_image_pdf:
                        # 使用改进的方法转换为图片式PDF，确保内容完整
                        self.log(f"正在将合同PDF转换为图片式PDF，确保内容完整...")
                        contract_pdf_final = self.convert_pdf_to_image_pdf(
                            contract_pdf_with_seal, 
                            dpi=200,  # 使用更高的DPI确保内容清晰
                            quality=85  # 使用更高的质量确保内容清晰
                        )
                        self.log(f"合同图片式PDF转换完成: {contract_pdf_final}")
                        
                        # 删除中间生成的PDF文件
                        if os.path.exists(contract_pdf_with_seal) and contract_pdf_with_seal != contract_pdf_final:
                            os.remove(contract_pdf_with_seal)
                            self.log(f"已删除中间PDF文件: {contract_pdf_with_seal}")
                        
                        return contract_file, contract_pdf_final
                    else:
                        # 返回带印章的PDF
                        return contract_file, contract_pdf_with_seal
                else:
                    self.log(f"未找到印章文件或印章文件不存在: {seal_file}")
                    # 返回不带印章的PDF
                    return contract_file, contract_pdf
            else:
                # 返回不带印章的PDF
                return contract_file, contract_pdf
        
        # 如果不需要转换为PDF，只返回Excel文件
        return contract_file
    
    def generate_quote_only(self, contract):
        """仅生成报价单文件"""
        # 生成报价单Excel文件
        quote_file = self._generate_quote_file(contract)
        
        # 如果需要转换为PDF
        if hasattr(contract, 'convert_to_pdf') and contract.convert_to_pdf:
            # 将Excel转换为PDF
            quote_pdf = self.convert_excel_to_pdf(quote_file)
            
            # 确保PDF为A4尺寸
            quote_pdf = self.ensure_a4_size_pdf(quote_pdf)
            
            # 如果需要添加印章
            if hasattr(contract, 'add_seal') and contract.add_seal:
                # 添加印章
                seal_file = self._get_seal_file(contract)
                if seal_file and os.path.exists(seal_file):
                    # 获取印章位置
                    position = 'right-bottom'
                    if hasattr(contract, 'seal_position'):
                        position = contract.seal_position
                    
                    # 获取签章文字
                    seal_text = None
                    if hasattr(contract, 'seal_text'):
                        seal_text = contract.seal_text
                    
                    # 添加印章到PDF
                    quote_pdf_with_seal = self.add_seal_to_pdf(
                        quote_pdf, 
                        seal_file, 
                        position,
                        seal_text=seal_text
                    )
                    
                    # 如果需要转换为图片式PDF
                    if hasattr(contract, 'convert_to_image_pdf') and contract.convert_to_image_pdf:
                        # 使用改进的方法转换为图片式PDF，确保内容完整
                        self.log(f"正在将报价单PDF转换为图片式PDF，确保内容完整...")
                        quote_pdf_final = self.convert_pdf_to_image_pdf(
                            quote_pdf_with_seal, 
                            dpi=200,  # 使用更高的DPI确保内容清晰
                            quality=85  # 使用更高的质量确保内容清晰
                        )
                        self.log(f"报价单图片式PDF转换完成: {quote_pdf_final}")
                        
                        # 删除中间生成的PDF文件
                        if os.path.exists(quote_pdf_with_seal) and quote_pdf_with_seal != quote_pdf_final:
                            os.remove(quote_pdf_with_seal)
                            self.log(f"已删除中间PDF文件: {quote_pdf_with_seal}")
                        
                        return quote_file, quote_pdf_final
                    else:
                        # 返回带印章的PDF
                        return quote_file, quote_pdf_with_seal
                else:
                    self.log(f"未找到印章文件或印章文件不存在: {seal_file}")
                    # 返回不带印章的PDF
                    return quote_file, quote_pdf
            else:
                # 返回不带印章的PDF
                return quote_file, quote_pdf
        
        # 如果不需要转换为PDF，只返回Excel文件
        return quote_file
    
    def _generate_contract_file(self, contract):
        """生成合同文件"""
        # 创建工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "购销合同"
        
        # 设置页面布局 - A4纸张适配
        ws.page_setup.paperSize = 9  # A4
        ws.page_setup.orientation = 'portrait'
        ws.page_setup.fitToPage = True
        
        # 设置列宽
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 24
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 10
        
        # 设置默认行高
        default_row_height = 25
        for i in range(1, 100):  # 预设100行的高度
            ws.row_dimensions[i].height = default_row_height
        
        # 定义样式
        title_font = Font(name='黑体', size=22, bold=True)
        header_font = Font(name='黑体', size=14, bold=True)
        subheader_font = Font(name='宋体', size=12, bold=True)
        normal_font = Font(name='宋体', size=12)
        normal_bold_font = Font(name='宋体', size=12, bold=True)
        small_font = Font(name='宋体', size=11)
        table_font = Font(name='宋体', size=12)
        
        # 定义边框样式
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 标题行
        row = 3
        ws.merge_cells(f'A{row}:H{row}')
        cell = ws[f'A{row}']
        cell.value = "购销合同"
        cell.font = title_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[row].height = 45
        
        # 合同编号 - 右上角，添加三位随机数
        row = 6
        ws.merge_cells(f'F{row}:H{row}')
        cell = ws[f'F{row}']
        # 添加三位随机数到合同编号
        contract_number = f"{contract.number}{contract.random_suffix}"
        cell.value = f"合同编号：{contract_number}"
        cell.font = normal_font
        cell.alignment = Alignment(horizontal='right', vertical='center')
        
        # 甲乙方信息
        row = 8
        ws[f'A{row}'].value = f"甲方（购货方）："
        ws[f'A{row}'].font = normal_font
        ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')
        ws.merge_cells(f'B{row}:H{row}')
        ws[f'B{row}'].value = contract.customer.name
        ws[f'B{row}'].font = normal_font
        ws[f'B{row}'].alignment = Alignment(horizontal='left', vertical='center')
        
        row += 1
        ws[f'A{row}'].value = f"乙方（销货方）："
        ws[f'A{row}'].font = normal_font
        ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')
        ws.merge_cells(f'B{row}:H{row}')
        ws[f'B{row}'].value = contract.company.get('name', '')
        ws[f'B{row}'].font = normal_font
        ws[f'B{row}'].alignment = Alignment(horizontal='left', vertical='center')
        
        # 合同依据说明
        row += 2
        ws.merge_cells(f'A{row}:H{row}')
        cell = ws[f'A{row}']
        cell.value = "根据《中华人民共和国民法典》及有关法律规定，经双方友好协商，就产品购销事宜达成以下协议："
        cell.font = normal_font
        cell.alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[row].height = 25
        
        # 空行
        row += 1
        
        # 商品清单标题
        row += 1
        ws.merge_cells(f'A{row}:B{row}')
        cell = ws[f'A{row}']
        cell.value = "一、商品清单："
        cell.font = subheader_font
        cell.alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[row].height = 25
        
        # 商品表头
        row += 1
        headers = ["序号", "商品名称", "规格型号", "单位", "数量", "单价(元)", "金额(元)", "备注"]
        for i, header in enumerate(headers):
            col = chr(ord('A') + i)
            ws[f'{col}{row}'].value = header
            ws[f'{col}{row}'].font = table_font
            ws[f'{col}{row}'].alignment = Alignment(horizontal='center', vertical='center')
            ws[f'{col}{row}'].border = thin_border
        ws.row_dimensions[row].height = 24
        
        # 商品数据
        start_row = row + 1
        for i, item in enumerate(contract.items):
            row = start_row + i
            # 序号
            ws[f'A{row}'].value = i + 1
            ws[f'A{row}'].font = table_font
            ws[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center')
            ws[f'A{row}'].border = thin_border
            
            # 商品名称
            ws[f'B{row}'].value = item.name
            ws[f'B{row}'].font = table_font
            ws[f'B{row}'].alignment = Alignment(horizontal='center', vertical='center')
            ws[f'B{row}'].border = thin_border
            
            # 规格型号
            ws[f'C{row}'].value = item.model
            ws[f'C{row}'].font = table_font
            ws[f'C{row}'].alignment = Alignment(horizontal='center', vertical='center')
            ws[f'C{row}'].border = thin_border
            
            # 单位
            ws[f'D{row}'].value = item.unit
            ws[f'D{row}'].font = table_font
            ws[f'D{row}'].alignment = Alignment(horizontal='center', vertical='center')
            ws[f'D{row}'].border = thin_border
            
            # 数量
            ws[f'E{row}'].value = item.quantity
            ws[f'E{row}'].font = table_font
            ws[f'E{row}'].alignment = Alignment(horizontal='center', vertical='center')
            ws[f'E{row}'].border = thin_border
            
            # 单价
            ws[f'F{row}'].value = item.price
            ws[f'F{row}'].font = table_font
            ws[f'F{row}'].alignment = Alignment(horizontal='center', vertical='center')
            ws[f'F{row}'].border = thin_border
            
            # 金额
            ws[f'G{row}'].value = f"=ROUND(E{row}*F{row}, 2)"  # 数量 * 单价，保留两位小数
            ws[f'G{row}'].font = table_font
            ws[f'G{row}'].alignment = Alignment(horizontal='center', vertical='center')
            ws[f'G{row}'].border = thin_border
            
            # 备注
            ws[f'H{row}'].value = ""
            ws[f'H{row}'].font = table_font
            ws[f'H{row}'].alignment = Alignment(horizontal='center', vertical='center')
            ws[f'H{row}'].border = thin_border
            
            # 根据内容长度调整行高
            name_length = len(str(item.name))
            model_length = len(str(item.model))
            if name_length > 10 or model_length > 15:
                ws.row_dimensions[row].height = 28
        
        # 设备小计
        row = start_row + len(contract.items)
        ws.merge_cells(f'A{row}:F{row}')
        ws[f'A{row}'].value = "设备小计"
        ws[f'A{row}'].font = table_font
        ws[f'A{row}'].alignment = Alignment(horizontal='right', vertical='center')
        
        # 为合并单元格设置完整边框
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            ws[f'{col}{row}'].border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        ws[f'G{row}'].value = f"=ROUND(SUM(G{start_row}:G{row-1}), 2)"  # 计算小计，保留两位小数
        ws[f'G{row}'].font = table_font
        ws[f'G{row}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'G{row}'].border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        ws[f'H{row}'].border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 技术服务费（合同）
        row += 1
        ws.merge_cells(f'A{row}:F{row}')
        ws[f'A{row}'].value = "技术服务费"
        ws[f'A{row}'].font = normal_font
        ws[f'A{row}'].alignment = Alignment(horizontal='right', vertical='center')
        
        # 为合并单元格设置完整边框
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            ws[f'{col}{row}'].border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        # 使用合同对象中的服务费参数
        if hasattr(contract, 'service_fee_enabled') and not contract.service_fee_enabled:
            # 如果禁用了技术服务费，设置为0
            ws[f'G{row}'].value = 0
        else:
            # 如果启用了技术服务费，使用费率和最低服务费计算
            service_fee_rate = contract.service_fee_rate if hasattr(contract, 'service_fee_rate') else 0.1
            min_service_fee = contract.min_service_fee if hasattr(contract, 'min_service_fee') else 1500
            ws[f'G{row}'].value = f"=MAX(ROUND(G{row-1}*{service_fee_rate}, 2), {min_service_fee})"
        
        ws[f'G{row}'].font = normal_font
        ws[f'G{row}'].alignment = Alignment(horizontal='center', vertical='center')
        
        # 设置备注列的完整边框
        ws[f'H{row}'].border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 合计
        row += 1
        ws.merge_cells(f'A{row}:F{row}')
        ws[f'A{row}'].value = "合计金额"
        ws[f'A{row}'].font = normal_bold_font
        ws[f'A{row}'].alignment = Alignment(horizontal='right', vertical='center')
        
        # 为合并单元格设置完整边框
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            ws[f'{col}{row}'].border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        ws[f'G{row}'].value = f"=ROUND(G{row-2}+G{row-1}, 2)"  # 小计 + 技术服务费，保留两位小数
        ws[f'G{row}'].font = normal_bold_font
        ws[f'G{row}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'G{row}'].border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        ws[f'H{row}'].border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 合同条款
        row += 2
        terms = [
            f"二、付款方式：{contract.payment_method}",
            f"三、交货时间：{contract.delivery_date}内交货。",
            "四、运输方式：物流（陆运），销货方负责邮费。",
            "五、质保条款：1年保修，维修邮寄费用各自承担。",
            "六、争议解决：甲乙双方协商，协商不成由乙方所在地仲裁机构或法院解决。",
            "七、合同生效：甲乙双方签字盖章后合同生效，合同一式两份，传真/扫描件有效。"
        ]
        
        for i, term in enumerate(terms):
            ws.merge_cells(f'A{row}:H{row}')
            # 所有条款使用普通字体，不加粗
            ws[f'A{row}'].value = term
            ws[f'A{row}'].font = normal_font
            ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')
            ws.row_dimensions[row].height = 24
            row += 1
        
        # 空白行 - 动态计算：12减去商品行数
        blank_rows = max(2, 12 - len(contract.items))  # 至少保留2行空白
        row += blank_rows
        
        # 签章区域
        ws.merge_cells(f'A{row}:D{row}')
        ws[f'A{row}'].value = f"甲方（盖章）：{contract.customer.name}"
        ws[f'A{row}'].font = normal_font
        ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')
        
        ws.merge_cells(f'E{row}:H{row}')
        ws[f'E{row}'].value = f"乙方（盖章）：{contract.company.get('name', '')}"
        ws[f'E{row}'].font = normal_font
        ws[f'E{row}'].alignment = Alignment(horizontal='left', vertical='center')
        
        # 甲方详细信息
        row += 1
        ws.merge_cells(f'A{row}:D{row}')
        ws[f'A{row}'].value = f"税号：{contract.customer.tax_id}"
        ws[f'A{row}'].font = small_font
        ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')
        
        ws.merge_cells(f'E{row}:H{row}')
        ws[f'E{row}'].value = f"税号：{contract.company.get('tax_id', '')}"
        ws[f'E{row}'].font = small_font
        ws[f'E{row}'].alignment = Alignment(horizontal='left', vertical='center')
        
        row += 1
        ws.merge_cells(f'A{row}:D{row}')
        ws[f'A{row}'].value = f"开户行：{contract.customer.bank_name}"
        ws[f'A{row}'].font = small_font
        ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')
        
        ws.merge_cells(f'E{row}:H{row}')
        ws[f'E{row}'].value = f"开户行：{contract.company.get('bank_name', '')}"
        ws[f'E{row}'].font = small_font
        ws[f'E{row}'].alignment = Alignment(horizontal='left', vertical='center')
        
        row += 1
        ws.merge_cells(f'A{row}:D{row}')
        ws[f'A{row}'].value = f"账号：{contract.customer.bank_account}"
        ws[f'A{row}'].font = small_font
        ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')
        
        ws.merge_cells(f'E{row}:H{row}')
        ws[f'E{row}'].value = f"账号：{contract.company.get('bank_account', '')}"
        ws[f'E{row}'].font = small_font
        ws[f'E{row}'].alignment = Alignment(horizontal='left', vertical='center')
        
        row += 1
        ws.merge_cells(f'A{row}:D{row}')
        ws[f'A{row}'].value = f"地址：{contract.customer.address}"
        ws[f'A{row}'].font = small_font
        ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')
        
        ws.merge_cells(f'E{row}:H{row}')
        ws[f'E{row}'].value = f"地址：{contract.company.get('address', '')}"
        ws[f'E{row}'].font = small_font
        ws[f'E{row}'].alignment = Alignment(horizontal='left', vertical='center')
        
        row += 1
        ws.merge_cells(f'A{row}:D{row}')
        ws[f'A{row}'].value = f"电话：{contract.customer.phone}"
        ws[f'A{row}'].font = small_font
        ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')
        
        ws.merge_cells(f'E{row}:H{row}')
        ws[f'E{row}'].value = f"电话：{contract.company.get('phone', '')}"
        ws[f'E{row}'].font = small_font
        ws[f'E{row}'].alignment = Alignment(horizontal='left', vertical='center')
        
        row += 1
        ws.merge_cells(f'A{row}:D{row}')
        ws[f'A{row}'].value = f"日期：{contract.sign_date}"
        ws[f'A{row}'].font = small_font
        ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')
        
        ws.merge_cells(f'E{row}:H{row}')
        ws[f'E{row}'].value = f"日期：{contract.sign_date}"
        ws[f'E{row}'].font = small_font
        ws[f'E{row}'].alignment = Alignment(horizontal='left', vertical='center')
        
        # 保存文件
        draft_suffix = "_草稿" if contract.is_draft else ""
        filename = f"{contract_number}{draft_suffix}.xlsx"
        filepath = os.path.join(self.output_dir, filename)
        wb.save(filepath)
        
        return filepath
    
    def _generate_quote_file(self, contract):
        """生成报价单文件"""
        # 创建工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "产品报价单"
        
        # 设置页面布局 - A4纸张适配
        ws.page_setup.paperSize = 9  # A4
        ws.page_setup.orientation = 'portrait'
        ws.page_setup.fitToPage = True
        
        # 设置列宽
        ws.column_dimensions['A'].width = 16   # 序号
        ws.column_dimensions['B'].width = 30  # 从22增加到30
        ws.column_dimensions['C'].width = 18  # 规格型号
        ws.column_dimensions['D'].width = 8   # 单位
        ws.column_dimensions['E'].width = 15   # 数量
        ws.column_dimensions['F'].width = 12  # 单价(元)
        ws.column_dimensions['G'].width = 15  # 金额(元)
        ws.column_dimensions['H'].width = 10  # 备注
        
        # 定义样式
        title_font = Font(name='黑体', size=20, bold=True)
        header_font = Font(name='黑体', size=12, bold=True)
        normal_font = Font(name='宋体', size=12)
        label_font = Font(name='宋体', size=12, bold=True)
        small_font = Font(name='宋体', size=10)
        
        # 定义边框样式
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 定义左右边框样式
        left_right_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin')
        )
        
        # 定义上下边框样式
        top_bottom_border = Border(
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 设置默认行高，确保文字不被隐藏
        default_row_height = 22
        for i in range(1, 100):  # 预设100行的高度
            ws.row_dimensions[i].height = default_row_height
        
        # 页眉分隔线
        row = 1
        ws.merge_cells(f'A{row}:H{row}')
        cell = ws[f'A{row}']
        # 移除页眉分隔线
        cell.border = Border()
        
        # 标题行 - 增加高度确保文字完全显示
        row = 2
        ws.merge_cells(f'A{row}:H{row}')
        cell = ws[f'A{row}']
        cell.value = "产品报价单"
        cell.font = title_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[row].height = 40
        
        # 上部分隔线
        row = 3
        ws.merge_cells(f'A{row}:H{row}')
        cell = ws[f'A{row}']
        # 移除上部分隔线
        cell.border = Border()
        
        # 公司信息区域 - 移除边框
        row = 4
        # 移除公司信息区块外框
        for r in range(4, 6):
            for c in range(ord('A'), ord('I')):
                col = chr(c)
                ws[f'{col}{r}'].border = Border()
        
        ws[f'A{row}'].value = "报价单位:"
        ws[f'A{row}'].font = normal_font
        ws[f'A{row}'].alignment = Alignment(horizontal='right', vertical='center')
        
        ws.merge_cells(f'B{row}:D{row}')
        ws[f'B{row}'].value = contract.company.get('name', '')
        ws[f'B{row}'].font = normal_font
        ws[f'B{row}'].alignment = Alignment(horizontal='left', vertical='center')
        
        ws[f'E{row}'].value = "报价单编号:"
        ws[f'E{row}'].font = normal_font
        ws[f'E{row}'].alignment = Alignment(horizontal='right', vertical='center')
        
        ws.merge_cells(f'F{row}:H{row}')
        # 使用BJ前缀加合同编号的后缀作为报价单编号，并添加三位随机数
        quote_number = f"BJ{contract.number[2:]}{contract.random_suffix}"
        ws[f'F{row}'].value = quote_number
        ws[f'F{row}'].font = normal_font
        ws[f'F{row}'].alignment = Alignment(horizontal='left', vertical='center')
        
        row = 5
        ws[f'A{row}'].value = "地址:"
        ws[f'A{row}'].font = normal_font
        ws[f'A{row}'].alignment = Alignment(horizontal='right', vertical='center')
        
        ws.merge_cells(f'B{row}:D{row}')
        ws[f'B{row}'].value = contract.company.get('address', '')
        ws[f'B{row}'].font = normal_font
        ws[f'B{row}'].alignment = Alignment(horizontal='left', vertical='center')
        
        ws[f'E{row}'].value = "报价日期:"
        ws[f'E{row}'].font = normal_font
        ws[f'E{row}'].alignment = Alignment(horizontal='right', vertical='center')
        
        ws.merge_cells(f'F{row}:H{row}')
        ws[f'F{row}'].value = contract.sign_date
        ws[f'F{row}'].font = normal_font
        ws[f'F{row}'].alignment = Alignment(horizontal='left', vertical='center')
        
        row = 6
        ws[f'A{row}'].value = "联系电话:"
        ws[f'A{row}'].font = normal_font
        ws[f'A{row}'].alignment = Alignment(horizontal='right', vertical='center')
        
        ws.merge_cells(f'B{row}:D{row}')
        ws[f'B{row}'].value = contract.company.get('phone', '')
        ws[f'B{row}'].font = normal_font
        ws[f'B{row}'].alignment = Alignment(horizontal='left', vertical='center')
        
        # 清空E6单元格内容
        ws[f'E{row}'].value = ""
        ws[f'E{row}'].font = normal_font
        ws[f'E{row}'].alignment = Alignment(horizontal='right', vertical='center')
        
        ws.merge_cells(f'F{row}:H{row}')
        # 清空F6单元格内容
        ws[f'F{row}'].value = ""
        ws[f'F{row}'].font = normal_font
        ws[f'F{row}'].alignment = Alignment(horizontal='left', vertical='center')
        
        # 空行分隔
        row = 7
        ws.merge_cells(f'A{row}:H{row}')
        
        # 客户信息区域 - 移除边框
        row = 8
        # 移除客户信息区块外框
        for r in range(8, 10):
            for c in range(ord('A'), ord('I')):
                col = chr(c)
                ws[f'{col}{r}'].border = Border()
        
        ws[f'A{row}'].value = "客户名称:"
        ws[f'A{row}'].font = normal_font
        ws[f'A{row}'].alignment = Alignment(horizontal='right', vertical='center')
        
        ws.merge_cells(f'B{row}:H{row}')
        ws[f'B{row}'].value = contract.customer.name
        ws[f'B{row}'].font = normal_font
        ws[f'B{row}'].alignment = Alignment(horizontal='left', vertical='center')
        
        row = 9
        ws[f'A{row}'].value = "联系电话:"
        ws[f'A{row}'].font = normal_font
        ws[f'A{row}'].alignment = Alignment(horizontal='right', vertical='center')
        
        ws.merge_cells(f'B{row}:H{row}')
        ws[f'B{row}'].value = contract.customer.phone
        ws[f'B{row}'].font = normal_font
        ws[f'B{row}'].alignment = Alignment(horizontal='left', vertical='center')
        
        # 空行分隔
        row = 10
        ws.merge_cells(f'A{row}:H{row}')
        
        # 问候语区域 - 移除边框
        row = 11
        ws.merge_cells(f'A{row}:H{row}')
        greeting = f"尊敬的客户，感谢您对我们的信任。以下是我们根据您的需求提供的产品报价，如有任何疑问，请随时联系我们。"
        ws[f'A{row}'].value = greeting
        ws[f'A{row}'].font = normal_font
        ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)  # 启用文本换行
        ws.row_dimensions[row].height = 35  # 增加问候语行高度
        
        # 移除问候语边框
        for col in range(ord('A'), ord('I')):
            ws[f'{chr(col)}{row}'].border = Border()
        
        # 空行分隔
        row = 12
        ws.merge_cells(f'A{row}:H{row}')
        
        # 产品报价明细标题
        row = 13
        ws.merge_cells(f'A{row}:H{row}')
        ws[f'A{row}'].value = "产品报价明细"
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[row].height = 25  # 增加标题行高度
        
        # 为标题添加完整边框
        for col in range(ord('A'), ord('I')):
            ws[f'{chr(col)}{row}'].border = thin_border
        
        # 商品表头
        row = 14
        headers = ["序号", "商品名称", "规格型号", "单位", "数量", "单价(元)", "金额(元)", "备注"]
        for i, header in enumerate(headers):
            col = chr(ord('A') + i)
            ws[f'{col}{row}'].value = header
            ws[f'{col}{row}'].font = normal_font
            ws[f'{col}{row}'].alignment = Alignment(horizontal='center', vertical='center')
            # 使用完整边框
            ws[f'{col}{row}'].border = thin_border
        
        # 商品数据
        start_row = row + 1
        for i, item in enumerate(contract.items):
            row = start_row + i
            
            # 序号
            ws[f'A{row}'].value = i + 1
            ws[f'A{row}'].font = normal_font
            ws[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center')
            ws[f'A{row}'].border = thin_border
            
            # 商品名称
            ws[f'B{row}'].value = item.name
            ws[f'B{row}'].font = normal_font
            ws[f'B{row}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws[f'B{row}'].border = thin_border
            
            # 规格型号
            ws[f'C{row}'].value = item.model
            ws[f'C{row}'].font = normal_font
            ws[f'C{row}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws[f'C{row}'].border = thin_border
            
            # 单位
            ws[f'D{row}'].value = item.unit
            ws[f'D{row}'].font = normal_font
            ws[f'D{row}'].alignment = Alignment(horizontal='center', vertical='center')
            ws[f'D{row}'].border = thin_border
            
            # 数量
            ws[f'E{row}'].value = item.quantity
            ws[f'E{row}'].font = normal_font
            ws[f'E{row}'].alignment = Alignment(horizontal='center', vertical='center')
            ws[f'E{row}'].border = thin_border
            
            # 单价
            ws[f'F{row}'].value = item.price
            ws[f'F{row}'].font = normal_font
            ws[f'F{row}'].alignment = Alignment(horizontal='center', vertical='center')
            ws[f'F{row}'].border = thin_border
            
            # 金额
            ws[f'G{row}'].value = f"=ROUND(E{row}*F{row}, 2)"
            ws[f'G{row}'].font = normal_font
            ws[f'G{row}'].alignment = Alignment(horizontal='center', vertical='center')
            ws[f'G{row}'].border = thin_border
            
            # 备注
            ws[f'H{row}'].value = ""
            ws[f'H{row}'].font = normal_font
            ws[f'H{row}'].alignment = Alignment(horizontal='center', vertical='center')
            ws[f'H{row}'].border = thin_border
            
            # 根据内容自动调整行高
            max_length = max(len(str(item.name)), len(str(item.model)))
            if max_length > 20:  # 如果内容较长，增加行高
                ws.row_dimensions[row].height = 28
        
        # 设备小计
        row = start_row + len(contract.items)
        ws.merge_cells(f'A{row}:F{row}')
        ws[f'A{row}'].value = "设备小计"
        ws[f'A{row}'].font = normal_font
        ws[f'A{row}'].alignment = Alignment(horizontal='right', vertical='center')
        
        # 为合并单元格设置完整边框
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            ws[f'{col}{row}'].border = thin_border
        
        ws[f'G{row}'].value = f"=ROUND(SUM(G{start_row}:G{row-1}), 2)"
        ws[f'G{row}'].font = normal_font
        ws[f'G{row}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'G{row}'].border = thin_border
        
        ws[f'H{row}'].border = thin_border
        
        # 技术服务费
        row += 1
        ws.merge_cells(f'A{row}:F{row}')
        ws[f'A{row}'].value = "技术服务费"
        ws[f'A{row}'].font = normal_font
        ws[f'A{row}'].alignment = Alignment(horizontal='right', vertical='center')
        
        # 为合并单元格设置完整边框
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            ws[f'{col}{row}'].border = thin_border
        
        if hasattr(contract, 'service_fee_enabled') and not contract.service_fee_enabled:
            ws[f'G{row}'].value = 0
        else:
            service_fee_rate = contract.service_fee_rate if hasattr(contract, 'service_fee_rate') else 0.1
            min_service_fee = contract.min_service_fee if hasattr(contract, 'min_service_fee') else 1500
            ws[f'G{row}'].value = f"=MAX(ROUND(G{row-1}*{service_fee_rate}, 2), {min_service_fee})"
        
        ws[f'G{row}'].font = normal_font
        ws[f'G{row}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'G{row}'].border = thin_border
        
        ws[f'H{row}'].border = thin_border
        
        # 合计金额
        row += 1
        ws.merge_cells(f'A{row}:F{row}')
        ws[f'A{row}'].value = "合计金额"
        ws[f'A{row}'].font = label_font
        ws[f'A{row}'].alignment = Alignment(horizontal='right', vertical='center')
        
        # 为合并单元格设置完整边框
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            ws[f'{col}{row}'].border = thin_border
        
        ws[f'G{row}'].value = f"=ROUND(G{row-2}+G{row-1}, 2)"
        ws[f'G{row}'].font = label_font
        ws[f'G{row}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'G{row}'].border = thin_border
        
        ws[f'H{row}'].border = thin_border
        
        # 报价说明标题
        row += 1
        ws.merge_cells(f'A{row}:H{row}')
        ws[f'A{row}'].value = "报价说明"
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[row].height = 25  # 增加标题行高度
        
        # 移除报价说明标题边框和背景色
        for col in range(ord('A'), ord('I')):
            ws[f'{chr(col)}{row}'].border = Border()
        
        # 报价说明内容
        row += 1
        quotes = [
            f"1. 本报价单有效期为{contract.quote_valid_days if hasattr(contract, 'quote_valid_days') else 15}天;",
            f"2. 付款方式: {contract.payment_method};",
            f"3. 预计交货日期: {contract.delivery_date};",
            "4. 以上价格为含税价格;",
            "5. 运输方式及费用由乙方承担;",
            "6. 如有其他特殊要求, 请与我们联系。"
        ]
        
        first_row = row
        for i, quote in enumerate(quotes):
            ws.merge_cells(f'A{row}:H{row}')
            ws[f'A{row}'].value = quote
            ws[f'A{row}'].font = normal_font
            ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)  # 启用文本换行
            
            # 移除边框
            for col in range(ord('A'), ord('I')):
                ws[f'{chr(col)}{row}'].border = Border()
            
            row += 1
        
        # 添加页脚 - 移除边框
        row += 1
        ws.merge_cells(f'A{row}:H{row}')
        ws[f'A{row}'].value = f"© {datetime.datetime.now().year} {contract.company.get('name', '')} - 专业服务，值得信赖"
        ws[f'A{row}'].font = small_font
        ws[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center')
        # 移除页脚边框
        ws[f'A{row}'].border = Border()
        
        # 保存文件
        filename = f"报价单-{contract.number}{contract.random_suffix}.xlsx"
        filepath = os.path.join(self.output_dir, filename)
        wb.save(filepath)
        
        return filepath
    
    def ensure_a4_size_pdf(self, pdf_file, output_pdf=None):
        """
        确保PDF文件为A4尺寸
        
        参数:
            pdf_file (str): PDF文件路径
            output_pdf (str, optional): 输出PDF文件路径，如果为None则覆盖原文件
            
        返回:
            str: 处理后的PDF文件路径
        """
        # 如果输出文件路径为None，则使用临时文件，之后再替换原文件
        is_same_file = False
        if output_pdf is None:
            is_same_file = True
            output_pdf = os.path.splitext(pdf_file)[0] + "_temp_a4.pdf"
        
        self.logger.info(f"开始调整PDF为A4尺寸: {pdf_file}")
        
        try:
            # 打开PDF文件
            pdf_document = fitz.open(pdf_file)
            
            # A4纸张尺寸（点，1英寸=72点）
            # A4尺寸为210mm x 297mm，转换为点为595 x 842
            a4_width = 595
            a4_height = 842
            
            # 创建一个新的PDF文档用于输出
            output_document = fitz.open()
            
            # 遍历所有页面
            for page_num in range(len(pdf_document)):
                page = pdf_document[page_num]
                
                # 获取页面尺寸
                page_width = page.rect.width
                page_height = page.rect.height
                
                # 创建一个新的A4尺寸页面
                new_page = output_document.new_page(width=a4_width, height=a4_height)
                
                # 检查页面尺寸是否需要调整
                if abs(page_width - a4_width) > 1 or abs(page_height - a4_height) > 1:
                    self.logger.info(f"页面尺寸不是A4，正在调整: {page_width}x{page_height} -> {a4_width}x{a4_height}")
                    
                    # 计算缩放比例
                    scale_x = a4_width / page_width
                    scale_y = a4_height / page_height
                    scale = min(scale_x, scale_y)  # 使用较小的缩放比例以保持纵横比
                    
                    # 计算居中位置
                    center_x = (a4_width - page_width * scale) / 2
                    center_y = (a4_height - page_height * scale) / 2
                    
                    # 将原页面内容绘制到新页面
                    new_page.show_pdf_page(
                        fitz.Rect(center_x, center_y, center_x + page_width * scale, center_y + page_height * scale),
                        pdf_document,
                        page_num
                    )
                else:
                    # 如果尺寸已经是A4，直接复制页面内容
                    new_page.show_pdf_page(
                        new_page.rect,
                        pdf_document,
                        page_num
                    )
            
            # 保存修改后的PDF
            output_document.save(output_pdf)
            output_document.close()
            pdf_document.close()
            
            # 如果是覆盖原文件，则在保存完成后替换原文件
            if is_same_file:
                # 确保关闭文件后再替换
                if os.path.exists(pdf_file):
                    os.remove(pdf_file)
                os.rename(output_pdf, pdf_file)
                output_pdf = pdf_file
            
            self.logger.info(f"成功调整PDF为A4尺寸: {output_pdf}")
            return output_pdf
            
        except Exception as e:
            self.logger.error(f"调整PDF为A4尺寸时出错: {str(e)}")
            raise 