#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
Excel工具模块
提供通用的Excel处理功能
"""

import os
from pathlib import Path
from typing import List, Dict, Any, Callable, Optional
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import (
    Font, Alignment, PatternFill, Border, Side,
    NamedStyle
)

class ExcelUtils:
    """Excel工具类"""
    
    @staticmethod
    def create_template(
        file_path: str,
        headers: List[str],
        header_display_names: List[str],
        required_fields: List[str] = None,
        example_data: List[Dict[str, Any]] = None
    ) -> str:
        """
        创建Excel模板文件
        
        Args:
            file_path: 模板文件路径
            headers: 表头字段列表
            header_display_names: 表头显示名称列表
            required_fields: 必填字段列表
            example_data: 示例数据列表
            
        Returns:
            str: 创建的模板文件路径
        """
        try:
            # 创建工作簿
            wb = Workbook()
            ws = wb.active
            
            # 创建表头样式
            header_style = NamedStyle(name='header_style')
            header_style.font = Font(bold=True, color='FFFFFF')
            header_style.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_style.alignment = Alignment(horizontal='center', vertical='center')
            header_style.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # 添加表头
            for idx, (header, display_name) in enumerate(zip(headers, header_display_names), 1):
                cell = ws.cell(row=1, column=idx, value=display_name)
                cell.style = header_style
                
                # 如果是必填字段，添加红色星号
                if required_fields and header in required_fields:
                    cell.value = f"{display_name}*"
            
            # 添加示例数据
            if example_data:
                for row_idx, data in enumerate(example_data, 2):
                    for col_idx, header in enumerate(headers, 1):
                        cell = ws.cell(row=row_idx, column=col_idx, value=data.get(header, ''))
                        cell.alignment = Alignment(horizontal='left', vertical='center')
            
            # 调整列宽
            for idx, header in enumerate(header_display_names, 1):
                column_letter = get_column_letter(idx)
                ws.column_dimensions[column_letter].width = max(len(str(header)) + 4, 15)
            
            # 冻结首行
            ws.freeze_panes = 'A2'
            
            # 保存模板文件
            wb.save(file_path)
            return file_path
            
        except Exception as e:
            raise Exception(f"创建模板文件失败：{str(e)}")
    
    @staticmethod
    def import_data(
        file_path: str,
        headers: List[str],
        header_display_names: List[str],
        required_fields: List[str],
        row_validator: Callable[[Dict[str, Any], int], Optional[str]],
        row_processor: Callable[[Dict[str, Any]], Any],
        progress_callback: Callable[[int, int], None] = None
    ) -> tuple[List[Any], List[str]]:
        """
        从Excel文件导入数据
        
        Args:
            file_path: Excel文件路径
            headers: 表头字段列表
            header_display_names: 表头显示名称列表
            required_fields: 必填字段列表
            row_validator: 行数据验证函数
            row_processor: 行数据处理函数
            progress_callback: 进度回调函数
            
        Returns:
            tuple: (导入的数据列表, 错误信息列表)
        """
        try:
            # 读取Excel文件
            wb = load_workbook(file_path, read_only=True)
            ws = wb.active
            
            # 获取表头
            file_headers = [str(cell.value).replace('*', '').strip() if cell.value else '' for cell in ws[1]]
            
            # 创建显示名称到字段名的映射和字段名到显示名称的映射
            header_display_map = dict(zip(header_display_names, headers))
            header_internal_map = dict(zip(headers, header_display_names))
            
            # 将文件表头转换为内部字段名
            internal_headers = []
            for header in file_headers:
                # 尝试直接匹配显示名称
                if header in header_display_map:
                    internal_headers.append(header_display_map[header])
                # 尝试匹配内部字段名
                elif header in headers:
                    internal_headers.append(header)
                else:
                    internal_headers.append(header)
            
            # 验证必要的列是否存在
            missing_columns = []
            for field in required_fields:
                if field not in internal_headers:
                    # 获取字段对应的显示名称
                    display_name = header_internal_map.get(field, field)
                    missing_columns.append(display_name)
            
            if missing_columns:
                raise Exception(
                    f"导入文件缺少必要的列：{', '.join(missing_columns)}\n"
                    "请使用正确的模板文件重新导入！"
                )
            
            # 创建列名到索引的映射
            header_map = {header: idx for idx, header in enumerate(internal_headers)}
            
            # 获取总行数（不包括表头）
            total_rows = sum(1 for _ in ws.iter_rows(min_row=2)) - 1
            
            # 导入数据
            imported_data = []
            validation_errors = []
            
            # 从第二行开始读取数据
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
                try:
                    # 创建行数据字典
                    row_data = {}
                    for header in headers:
                        if header in header_map:
                            value = row[header_map[header]]
                            row_data[header] = str(value).strip() if value is not None else ""
                    
                    # 验证数据
                    error = row_validator(row_data, row_idx)
                    if error:
                        validation_errors.append(f"第 {row_idx} 行：{error}")
                        continue
                    
                    # 处理数据
                    processed_data = row_processor(row_data)
                    if processed_data:
                        imported_data.append(processed_data)
                    
                    # 更新进度
                    if progress_callback:
                        progress_callback(row_idx - 1, total_rows)
                
                except Exception as e:
                    validation_errors.append(f"第 {row_idx} 行：数据格式错误 - {str(e)}")
            
            return imported_data, validation_errors
            
        except Exception as e:
            raise Exception(f"导入数据失败：{str(e)}")
    
    @staticmethod
    def export_data(
        file_path: str,
        headers: List[str],
        header_display_names: List[str],
        data: List[Any],
        row_formatter: Callable[[Any], List[Any]],
        progress_callback: Callable[[int, int], None] = None
    ) -> None:
        """
        导出数据到Excel文件
        
        Args:
            file_path: 导出文件路径
            headers: 表头字段列表
            header_display_names: 表头显示名称列表
            data: 要导出的数据列表
            row_formatter: 行数据格式化函数
            progress_callback: 进度回调函数
        """
        try:
            # 创建工作簿
            wb = Workbook()
            ws = wb.active
            
            # 创建表头样式
            header_style = NamedStyle(name='header_style')
            header_style.font = Font(bold=True, color='FFFFFF')
            header_style.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_style.alignment = Alignment(horizontal='center', vertical='center')
            header_style.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # 添加表头
            for idx, display_name in enumerate(header_display_names, 1):
                cell = ws.cell(row=1, column=idx, value=display_name)
                cell.style = header_style
            
            # 添加数据
            total_rows = len(data)
            for row_idx, item in enumerate(data, 2):
                # 格式化行数据
                row_data = row_formatter(item)
                
                # 写入数据
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                    cell.border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                
                # 更新进度
                if progress_callback:
                    progress_callback(row_idx - 1, total_rows)
            
            # 调整列宽
            for idx, header in enumerate(header_display_names, 1):
                # 计算列的最大宽度
                max_length = len(str(header))
                column_letter = get_column_letter(idx)
                
                for row_idx in range(2, ws.max_row + 1):
                    cell_value = ws.cell(row=row_idx, column=idx).value
                    if cell_value:
                        max_length = max(max_length, len(str(cell_value)))
                
                ws.column_dimensions[column_letter].width = max_length + 4
            
            # 冻结首行
            ws.freeze_panes = 'A2'
            
            # 保存文件
            wb.save(file_path)
            
        except Exception as e:
            raise Exception(f"导出数据失败：{str(e)}") 